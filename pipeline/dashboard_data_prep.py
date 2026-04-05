"""
PMO Governance Dashboard — Dashboard Data Prep & Excel Report Generator
Author: Anandi Mahure
Description: Generates Power BI-ready CSV outputs and a formatted Excel
governance pack with RAG summary, budget variance, and milestone tracking.
"""

import pandas as pd
import numpy as np
import os
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

os.makedirs("outputs", exist_ok=True)

NAVY   = "1F4E79"
RED    = "FFC7CE"
AMBER  = "FFEB9C"
GREEN  = "C6EFCE"
WHITE  = "FFFFFF"
GREY   = "F2F2F2"


def _thin_border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)


def build_rag_summary_sheet(ws, df: pd.DataFrame, run_date: str):
    ws.title = "01 RAG Summary"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B1:G1")
    ws["B1"].value = "PMO GOVERNANCE DASHBOARD — WEEKLY PACK"
    ws["B1"].font = Font(name="Arial", bold=True, size=14, color=WHITE)
    ws["B1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    ws["B2"].value = f"Run Date: {run_date}   |   Portfolio: {len(df)} Projects"
    ws["B2"].font = Font(name="Arial", size=10, italic=True, color="666666")
    ws.row_dimensions[2].height = 18

    rag_counts = df["rag_status"].value_counts()
    kpis = [
        ("C4", "Total Projects",   str(len(df)),                        NAVY),
        ("D4", "🔴 RED",           str(rag_counts.get("RED", 0)),        "FF0000"),
        ("E4", "🟡 AMBER",         str(rag_counts.get("AMBER", 0)),      "FFA500"),
        ("F4", "🟢 GREEN",         str(rag_counts.get("GREEN", 0)),      "00B050"),
        ("G4", "Portfolio Variance",
         f"{round((df['actual_spend'].sum()-df['total_budget'].sum())/df['total_budget'].sum()*100,1):+.1f}%",
         NAVY),
    ]
    for ref, label, value, colour in kpis:
        col = ref[0]
        row = int(ref[1])
        label_cell = ws[f"{col}{row}"]
        label_cell.value = label
        label_cell.font = Font(name="Arial", bold=True, size=9, color=WHITE)
        label_cell.fill = PatternFill("solid", fgColor=NAVY)
        label_cell.alignment = Alignment(horizontal="center")

        val_cell = ws[f"{col}{row+1}"]
        val_cell.value = value
        val_cell.font = Font(name="Arial", bold=True, size=20, color=colour)
        val_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row+1].height = 40

    # RAG table
    headers = ["Project ID", "Project Name", "Owner", "RAG",
               "Budget Variance %", "Schedule (Days)", "Risk Score", "Status"]
    for col_idx, h in enumerate(headers, start=2):
        cell = ws.cell(row=7, column=col_idx, value=h)
        cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center")

    rag_fill = {"RED": RED, "AMBER": AMBER, "GREEN": GREEN}
    for row_idx, (_, row) in enumerate(df.iterrows(), start=8):
        fill_colour = rag_fill.get(row["rag_status"], WHITE)
        values = [row["project_id"], row["project_name"], row["project_owner"],
                  row["rag_status"], f"{row['budget_variance_pct']:+.1f}%",
                  row["schedule_variance_days"], row["risk_score"], row["status"]]
        for col_idx, val in enumerate(values, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=9)
            cell.fill = PatternFill("solid", fgColor=fill_colour)
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal="center")

    for col, width in zip("BCDEFGHI", [12, 28, 14, 10, 16, 16, 12, 14]):
        ws.column_dimensions[col].width = width

    log.info("  Sheet 01 RAG Summary — built")


def main():
    log.info("=" * 55)
    log.info("PMO GOVERNANCE — DASHBOARD DATA PREP STAGE")
    log.info("=" * 55)

    df = pd.read_csv("data/projects.csv")
    run_date = datetime.now().strftime("%d %B %Y %H:%M")

    wb = Workbook()
    wb.remove(wb.active)
    build_rag_summary_sheet(wb.create_sheet(), df, run_date)

    output_path = "outputs/weekly_governance_pack.xlsx"
    wb.save(output_path)
    log.info(f"\nGovernance pack saved: {output_path}")

    # Power BI ready CSVs
    df.to_csv("outputs/pbi_projects.csv", index=False)
    log.info("Power BI dataset saved: outputs/pbi_projects.csv")


if __name__ == "__main__":
    main()
